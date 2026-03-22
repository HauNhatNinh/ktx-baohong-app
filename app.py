from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file
from flask_cors import CORS
import mysql.connector
from mysql.connector import Error
import os
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import cloudinary
import cloudinary.uploader
import cloudinary.api
from dotenv import load_dotenv
import pandas as pd
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import random
import string
import threading

load_dotenv()

# Cau hinh Cloudinary
cloudinary.config(
    cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key=os.environ.get('CLOUDINARY_API_KEY'),
    api_secret=os.environ.get('CLOUDINARY_API_SECRET'),
    secure=True
)

app = Flask(__name__)
app.secret_key = 'ktx_baohong_secret_key_2024'
CORS(app)

# Cau hinh upload anh
THU_MUC_UPLOAD = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
if not os.path.exists(THU_MUC_UPLOAD):
    os.makedirs(THU_MUC_UPLOAD)
app.config['THU_MUC_UPLOAD'] = THU_MUC_UPLOAD
DUOI_FILE_CHO_PHEP = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def kiem_tra_duoi_file(ten_file):
    return '.' in ten_file and ten_file.rsplit('.', 1)[1].lower() in DUOI_FILE_CHO_PHEP

# ============================
# KET NOI DATABASE
# ============================
def ket_noi_db():
    try:
        conn = mysql.connector.connect(
            host=os.environ.get('DB_HOST', 'mysql-ktxbaohong-ktxbaohong-2026.e.aivencloud.com'),
            port=int(os.environ.get('DB_PORT', 28894)),
            user=os.environ.get('DB_USER', 'avnadmin'),
            password=os.environ.get('DB_PASSWORD', ''),
            database=os.environ.get('DB_NAME', 'defaultdb'),
            charset='utf8mb4',
            ssl_disabled=False
        )
        cur = conn.cursor()
        cur.execute("SET time_zone = '+07:00'")
        cur.close()
        return conn
    except Error as e:
        print(f"Loi ket noi database: {e}")
        return None

# ============================
# TRANG CHINH
# ============================
@app.route('/')
def trang_chu():
    return render_template('login.html')

@app.route('/trang-sinh-vien')
def trang_sinh_vien():
    if 'ma_nguoidung' not in session or session.get('vai_tro') != 'sinhvien':
        return redirect('/')
    return render_template('student.html')

@app.route('/trang-quan-ly')
def trang_quan_ly():
    if 'ma_nguoidung' not in session or session.get('vai_tro') != 'quanly':
        return redirect('/')
    return render_template('manager.html')

@app.route('/trang-ky-thuat')
def trang_ky_thuat():
    if 'ma_nguoidung' not in session or session.get('vai_tro') != 'kythuat':
        return redirect('/')
    return render_template('technician.html')

@app.route('/trang-admin')
def trang_admin():
    if 'ma_nguoidung' not in session or session.get('vai_tro') != 'admin':
        return redirect('/')
    return render_template('admin.html')

@app.route('/uploads/<ten_file>')
def lay_file_upload(ten_file):
    from flask import send_from_directory
    return send_from_directory(app.config['THU_MUC_UPLOAD'], ten_file)

# ============================
# API XAC THUC
# ============================
@app.route('/api/dang-nhap', methods=['POST'])
def dang_nhap():
    du_lieu = request.get_json()
    tai_khoan = du_lieu.get('tai_khoan', '')
    mat_khau = du_lieu.get('mat_khau', '')
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Lỗi kết nối cơ sở dữ liệu'}), 500
    
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute(
            "SELECT * FROM nguoidung WHERE tai_khoan = %s",
            (tai_khoan,)
        )
        nguoi_dung = con_tro.fetchone()
        
        # Kiểm tra mật khẩu mã hoá (hash) thay vì so sánh chữ thô
        if not nguoi_dung or not check_password_hash(nguoi_dung['mat_khau'], mat_khau):
            return jsonify({'thanh_cong': False, 'thong_bao': 'Sai tài khoản hoặc mật khẩu'})
        
        if nguoi_dung['trang_thai'] == 'cho_duyet':
            return jsonify({'thanh_cong': False, 'thong_bao': 'Tài khoản đang chờ duyệt. Vui lòng liên hệ quản lý ký túc xá.'})
        
        if nguoi_dung['trang_thai'] == 'tu_choi':
            return jsonify({'thanh_cong': False, 'thong_bao': 'Yêu cầu đăng ký đã bị từ chối.'})
        
        session['ma_nguoidung'] = nguoi_dung['ma_nguoidung']
        session['tai_khoan'] = nguoi_dung['tai_khoan']
        session['vai_tro'] = nguoi_dung['vai_tro']
        session['ho_ten'] = nguoi_dung['ho_ten']
        
        duong_dan = {
            'sinhvien': '/trang-sinh-vien',
            'quanly': '/trang-quan-ly',
            'kythuat': '/trang-ky-thuat',
            'admin': '/trang-admin'
        }
        
        return jsonify({
            'thanh_cong': True,
            'thong_bao': 'Đăng nhập thành công',
            'vai_tro': nguoi_dung['vai_tro'],
            'duong_dan': duong_dan.get(nguoi_dung['vai_tro'], '/')
        })
    finally:
        conn.close()

@app.route('/api/dang-ky', methods=['POST'])
def dang_ky():
    # Ho tro ca JSON va FormData
    if request.is_json:
        du_lieu = request.get_json()
        tai_khoan = du_lieu.get('tai_khoan', '')
        mat_khau = du_lieu.get('mat_khau', '')
        ho_ten = du_lieu.get('ho_ten', '')
        email = du_lieu.get('email', '')
        lop = du_lieu.get('lop', '')
        khoa = du_lieu.get('khoa', '')
        ma_phong = du_lieu.get('ma_phong')
    else:
        tai_khoan = request.form.get('tai_khoan', '')
        mat_khau = request.form.get('mat_khau', '')
        ho_ten = request.form.get('ho_ten', '')
        email = request.form.get('email', '')
        lop = request.form.get('lop', '')
        khoa = request.form.get('khoa', '')
        ma_phong = request.form.get('ma_phong')

    if not all([tai_khoan, mat_khau, ho_ten, email, lop, khoa, ma_phong]):
        return jsonify({'thanh_cong': False, 'thong_bao': 'Vui lòng điền đầy đủ thông tin'})
    
    # Xu ly anh the sinh vien
    ten_file_anh = None
    if 'anh_the_sv' in request.files:
        file_anh = request.files['anh_the_sv']
        if file_anh and file_anh.filename and kiem_tra_duoi_file(file_anh.filename):
            if os.environ.get('CLOUDINARY_CLOUD_NAME'):
                try:
                    upload_result = cloudinary.uploader.upload(file_anh)
                    ten_file_anh = upload_result.get('secure_url')
                except Exception as e:
                    print(f"Loi upload Cloudinary (Anh the SV): {e}")
            else:
                ten_file = secure_filename(f"the_sv_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file_anh.filename}")
                file_anh.save(os.path.join(app.config['THU_MUC_UPLOAD'], ten_file))
                ten_file_anh = ten_file
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Lỗi kết nối cơ sở dữ liệu'}), 500
    
    try:
        con_tro = conn.cursor(dictionary=True)
        # Kiem tra tai khoan hoac email da ton tai
        con_tro.execute("SELECT * FROM nguoidung WHERE tai_khoan = %s", (tai_khoan,))
        user_tk = con_tro.fetchone()
        if user_tk:
            if user_tk['trang_thai'] != 'tu_choi':
                return jsonify({'thanh_cong': False, 'thong_bao': 'Mã số sinh viên đã được đăng ký'})
            
        con_tro.execute("SELECT * FROM nguoidung WHERE email = %s", (email,))
        user_email = con_tro.fetchone()
        if user_email:
            if user_email['trang_thai'] != 'tu_choi':
                return jsonify({'thanh_cong': False, 'thong_bao': 'Email này đã được sử dụng'})
        
        # Xoá dữ liệu cũ nếu bị từ chối trước đó để cho phép đăng ký lại
        con_tro.execute("DELETE FROM nguoidung WHERE (tai_khoan = %s OR email = %s) AND trang_thai = 'tu_choi'", (tai_khoan, email))
        
        hashed_password = generate_password_hash(mat_khau)
        con_tro.execute(
            """INSERT INTO nguoidung (tai_khoan, mat_khau, ho_ten, email, anh_the_sv, vai_tro, trang_thai, lop, khoa, ma_phong)
            VALUES (%s, %s, %s, %s, %s, 'sinhvien', 'cho_duyet', %s, %s, %s)""",
            (tai_khoan, hashed_password, ho_ten, email, ten_file_anh, lop, khoa, ma_phong)
        )
        conn.commit()
        
        # Gửi email thông báo ngay khi vừa đăng ký
        if email:
            tieu_de = "⏳ Tiếp nhận thông tin Đăng Ký Tài Khoản KTX"
            noi_dung = f'''
            <h3>Xin chào {ho_ten},</h3>
            <p>Hệ thống Ký túc xá đã nhận được yêu cầu đăng ký tài khoản của bạn với MSSV: <b>{tai_khoan}</b>.</p>
            <p>Tài khoản của bạn hiện đang ở trạng thái <b>Chờ Xét Duyệt</b>.</p>
            <p>Ban Quản Lý KTX sẽ kiểm tra thông tin và ảnh thẻ sinh viên của bạn. Khi được duyệt, chúng tôi sẽ gửi thêm một email thông báo nữa cho bạn.</p>
            <br>
            <p>Cảm ơn bạn.<br><b>Ban Quản Lý KTX</b></p>
            '''
            threading.Thread(target=gui_email_thong_bao, args=(email, tieu_de, noi_dung), daemon=True).start()
            
        return jsonify({'thanh_cong': True, 'thong_bao': 'Đã tiếp nhận phiếu đăng ký, vui lòng chờ quản lý xét duyệt tài khoản'})
    except Exception as e:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Có lỗi xảy ra, thử lại sau.'})
    finally:
        conn.close()

@app.route('/api/dang-xuat', methods=['POST'])
def dang_xuat():
    session.clear()
    return jsonify({'thanh_cong': True})

# ============================
# API LAY DANH SACH TOA NHA & PHONG
# ============================
@app.route('/api/danh-sach-toanha', methods=['GET'])
def danh_sach_toanha():
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("SELECT * FROM toanha ORDER BY ten_toanha")
        ds_toanha = con_tro.fetchall()
        return jsonify({'thanh_cong': True, 'du_lieu': ds_toanha})
    finally:
        conn.close()

@app.route('/api/danh-sach-phong/<int:ma_toanha>', methods=['GET'])
def danh_sach_phong(ma_toanha):
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("SELECT * FROM phong WHERE ma_toanha = %s ORDER BY ten_phong", (ma_toanha,))
        ds_phong = con_tro.fetchall()
        return jsonify({'thanh_cong': True, 'du_lieu': ds_phong})
    finally:
        conn.close()

# ============================
# API SINH VIEN
# ============================
@app.route('/api/sinhvien/thong-tin', methods=['GET'])
def lay_thong_tin_sv():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Chưa đăng nhập'}), 401
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("""
            SELECT nd.*, p.ten_phong, t.ten_toanha 
            FROM nguoidung nd
            LEFT JOIN phong p ON nd.ma_phong = p.ma_phong
            LEFT JOIN toanha t ON p.ma_toanha = t.ma_toanha
            WHERE nd.ma_nguoidung = %s
        """, (session['ma_nguoidung'],))
        thong_tin = con_tro.fetchone()
        if thong_tin:
            thong_tin.pop('mat_khau', None)
            if thong_tin.get('ngay_sinh'):
                thong_tin['ngay_sinh'] = thong_tin['ngay_sinh'].strftime('%Y-%m-%d')
            if thong_tin.get('ngay_tao'):
                thong_tin['ngay_tao'] = thong_tin['ngay_tao'].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify({'thanh_cong': True, 'du_lieu': thong_tin})
    finally:
        conn.close()

@app.route('/api/sinhvien/cap-nhat', methods=['PUT'])
def cap_nhat_sv():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    
    if request.is_json:
        du_lieu = request.get_json()
        ho_ten = du_lieu.get('ho_ten')
        lop = du_lieu.get('lop')
        khoa = du_lieu.get('khoa')
        so_dien_thoai = du_lieu.get('so_dien_thoai')
        ma_phong = du_lieu.get('ma_phong')
    else:
        ho_ten = request.form.get('ho_ten')
        lop = request.form.get('lop')
        khoa = request.form.get('khoa')
        so_dien_thoai = request.form.get('so_dien_thoai')
        ma_phong = request.form.get('ma_phong')
        if ma_phong == 'null' or not ma_phong or ma_phong == '':
            ma_phong = None
    
    ten_file_anh = None
    if 'anh_dai_dien' in request.files:
        file_anh = request.files['anh_dai_dien']
        if file_anh and file_anh.filename and kiem_tra_duoi_file(file_anh.filename):
            if os.environ.get('CLOUDINARY_CLOUD_NAME'):
                try:
                    upload_result = cloudinary.uploader.upload(file_anh)
                    ten_file_anh = upload_result.get('secure_url')
                except Exception as e:
                    print(f"Loi upload Cloudinary: {e}")
            else:
                ten_file = secure_filename(f"avatar_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file_anh.filename}")
                file_anh.save(os.path.join(app.config['THU_MUC_UPLOAD'], ten_file))
                ten_file_anh = ten_file
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor()
        if ten_file_anh:
            con_tro.execute("""
                UPDATE nguoidung SET ho_ten=%s, lop=%s, khoa=%s, so_dien_thoai=%s, ma_phong=%s, anh_dai_dien=%s WHERE ma_nguoidung=%s
            """, (ho_ten, lop, khoa, so_dien_thoai, ma_phong, ten_file_anh, session['ma_nguoidung']))
        else:
            con_tro.execute("""
                UPDATE nguoidung SET ho_ten=%s, lop=%s, khoa=%s, so_dien_thoai=%s, ma_phong=%s WHERE ma_nguoidung=%s
            """, (ho_ten, lop, khoa, so_dien_thoai, ma_phong, session['ma_nguoidung']))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Cập nhật thành công'})
    finally:
        conn.close()

@app.route('/api/sinhvien/thanh-vien-phong', methods=['GET'])
def thanh_vien_phong():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        # Lay ma phong cua sinh vien hien tai
        con_tro.execute("SELECT ma_phong FROM nguoidung WHERE ma_nguoidung = %s", (session['ma_nguoidung'],))
        sv = con_tro.fetchone()
        if not sv or not sv['ma_phong']:
            return jsonify({'thanh_cong': True, 'du_lieu': []})
        
        con_tro.execute("""
            SELECT ma_nguoidung, tai_khoan, ho_ten, lop, khoa, anh_dai_dien, so_dien_thoai
            FROM nguoidung 
            WHERE ma_phong = %s AND vai_tro = 'sinhvien' AND trang_thai = 'da_duyet'
            AND ma_nguoidung != %s
        """, (sv['ma_phong'], session['ma_nguoidung']))
        ds_thanh_vien = con_tro.fetchall()
        return jsonify({'thanh_cong': True, 'du_lieu': ds_thanh_vien})
    finally:
        conn.close()

@app.route('/api/sinhvien/tao-phieu', methods=['POST'])
def tao_phieu_bao_hong():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    
    ten_loi = request.form.get('ten_loi', '')
    mo_ta = request.form.get('mo_ta', '')
    muc_do = request.form.get('muc_do', 'thuong')
    sdt_lien_he = request.form.get('sdt_lien_he', '')
    
    # Ghep SDT vao mo ta de de dang thay ma khong can them cot DB
    if sdt_lien_he:
        mo_ta = f"[SĐT Liên Hệ: {sdt_lien_he}]\n{mo_ta}"
    
    if not ten_loi:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Vui lòng nhập tên lỗi hỏng'})
    
    # Xu ly upload anh
    ten_file_anh = ''
    if 'anh_minh_chung' in request.files:
        file_anh = request.files['anh_minh_chung']
        if file_anh and file_anh.filename and kiem_tra_duoi_file(file_anh.filename):
            if os.environ.get('CLOUDINARY_CLOUD_NAME'):
                try:
                    upload_result = cloudinary.uploader.upload(file_anh)
                    ten_file_anh = upload_result.get('secure_url')
                except Exception as e:
                    print(f"Loi upload Cloudinary: {e}")
            else:
                ten_file = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file_anh.filename}")
                file_anh.save(os.path.join(app.config['THU_MUC_UPLOAD'], ten_file))
                ten_file_anh = ten_file
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        # Lay thong tin phong va toa cua sinh vien
        con_tro.execute("""
            SELECT nd.ma_phong, p.ma_toanha 
            FROM nguoidung nd
            JOIN phong p ON nd.ma_phong = p.ma_phong
            WHERE nd.ma_nguoidung = %s
        """, (session['ma_nguoidung'],))
        thong_tin_sv = con_tro.fetchone()
        
        if not thong_tin_sv:
            return jsonify({'thanh_cong': False, 'thong_bao': 'Không tìm thấy thông tin phòng'})
        
        con_tro.execute("""
            INSERT INTO phieubaohong (ma_nguoidung, ma_phong, ma_toanha, ten_loi, mo_ta, anh_minh_chung, muc_do, trang_thai)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            session['ma_nguoidung'],
            thong_tin_sv['ma_phong'],
            thong_tin_sv['ma_toanha'],
            ten_loi, mo_ta, ten_file_anh, muc_do,
            'cho_xu_ly'
        ))
        ma_phieu = con_tro.lastrowid
        
        # Neu khan cap: tu dong gui den tat ca ky thuat vien
        if muc_do == 'khancap':
            con_tro.execute("SELECT ma_nguoidung FROM nguoidung WHERE vai_tro = 'kythuat' AND trang_thai = 'da_duyet'")
            ds_kythuat = con_tro.fetchall()
            for kt in ds_kythuat:
                # Tim quan ly cua toa nha
                con_tro.execute("""
                    SELECT ma_nguoidung FROM nguoidung 
                    WHERE vai_tro = 'quanly' AND ma_toanha_quanly = %s AND trang_thai = 'da_duyet'
                    LIMIT 1
                """, (thong_tin_sv['ma_toanha'],))
                quanly = con_tro.fetchone()
                ma_quanly = quanly['ma_nguoidung'] if quanly else session['ma_nguoidung']
                
                con_tro.execute("""
                    INSERT INTO phancong (ma_phieu, ma_kythuat, ma_quanly, trang_thai)
                    VALUES (%s, %s, %s, 'cho_tiep_nhan')
                """, (ma_phieu, kt['ma_nguoidung'], ma_quanly))
            
            # Cap nhat trang thai phieu thanh da phan cong
            con_tro.execute("UPDATE phieubaohong SET trang_thai = 'da_phan_cong' WHERE ma_phieu = %s", (ma_phieu,))
        
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Tạo phiếu báo hỏng thành công!'})
    finally:
        conn.close()

@app.route('/api/sinhvien/lich-su', methods=['GET'])
def lich_su_bao_hong():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("""
            SELECT pb.*, p.ten_phong, t.ten_toanha, n_tao.ho_ten AS nguoi_bao_cao, n_tao.tai_khoan AS mssv_nguoi_bao,
                   pc.anh_hoan_thanh, pc.ghi_chu, pc.ngay_cap_nhat AS thoi_gian_sua_xong, kt.ho_ten AS ten_ky_thuat, kt.so_dien_thoai AS sdt_ky_thuat
            FROM phieubaohong pb
            JOIN phong p ON pb.ma_phong = p.ma_phong
            JOIN toanha t ON pb.ma_toanha = t.ma_toanha
            LEFT JOIN nguoidung n_tao ON pb.ma_nguoidung = n_tao.ma_nguoidung
            LEFT JOIN phancong pc ON pc.ma_phieu = pb.ma_phieu AND pc.trang_thai IN ('da_tiep_nhan', 'hoan_thanh')
            LEFT JOIN nguoidung kt ON pc.ma_kythuat = kt.ma_nguoidung
            WHERE pb.ma_phong = (SELECT ma_phong FROM nguoidung WHERE ma_nguoidung = %s)
            ORDER BY pb.ngay_tao DESC
        """, (session['ma_nguoidung'],))
        ds_phieu = con_tro.fetchall()
        for phieu in ds_phieu:
            phieu['ngay_tao'] = phieu['ngay_tao'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_tao'] else ''
            phieu['ngay_cap_nhat'] = phieu['ngay_cap_nhat'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_cap_nhat'] else ''
            phieu['thoi_gian_sua_xong'] = phieu['thoi_gian_sua_xong'].strftime('%d/%m/%Y %H:%M') if phieu.get('thoi_gian_sua_xong') else ''
        return jsonify({'thanh_cong': True, 'du_lieu': ds_phieu})
    finally:
        conn.close()

# ============================
# API QUAN LY KTX
# ============================
@app.route('/api/quanly/thong-tin', methods=['GET'])
def lay_thong_tin_ql():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("""
            SELECT nd.*, t.ten_toanha 
            FROM nguoidung nd
            LEFT JOIN toanha t ON nd.ma_toanha_quanly = t.ma_toanha
            WHERE nd.ma_nguoidung = %s
        """, (session['ma_nguoidung'],))
        thong_tin = con_tro.fetchone()
        if thong_tin:
            thong_tin.pop('mat_khau', None)
            if thong_tin.get('ngay_sinh'):
                thong_tin['ngay_sinh'] = thong_tin['ngay_sinh'].strftime('%Y-%m-%d')
            if thong_tin.get('ngay_tao'):
                thong_tin['ngay_tao'] = thong_tin['ngay_tao'].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify({'thanh_cong': True, 'du_lieu': thong_tin})
    finally:
        conn.close()

@app.route('/api/quanly/cap-nhat', methods=['PUT'])
def cap_nhat_ql():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    du_lieu = request.get_json()
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor()
        con_tro.execute("""
            UPDATE nguoidung SET ho_ten=%s, ma_toanha_quanly=%s, ngay_sinh=%s, so_dien_thoai=%s
            WHERE ma_nguoidung=%s
        """, (
            du_lieu.get('ho_ten'),
            du_lieu.get('ma_toanha_quanly'),
            du_lieu.get('ngay_sinh'),
            du_lieu.get('so_dien_thoai'),
            session['ma_nguoidung']
        ))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Cập nhật thành công'})
    finally:
        conn.close()

@app.route('/api/quanly/danh-sach-phong', methods=['GET'])
def ql_danh_sach_phong():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        # Lay toa nha quan ly
        con_tro.execute("SELECT ma_toanha_quanly FROM nguoidung WHERE ma_nguoidung = %s", (session['ma_nguoidung'],))
        ql = con_tro.fetchone()
        ma_toanha = ql.get('ma_toanha_quanly') if ql else None
        
        if not ma_toanha:
            # Neu chua co toa, tra ve tat ca phong
            con_tro.execute("""
                SELECT p.*, t.ten_toanha FROM phong p
                JOIN toanha t ON p.ma_toanha = t.ma_toanha
                ORDER BY t.ten_toanha, p.ten_phong
            """)
        else:
            con_tro.execute("""
                SELECT p.*, t.ten_toanha FROM phong p
                JOIN toanha t ON p.ma_toanha = t.ma_toanha
                WHERE p.ma_toanha = %s
                ORDER BY p.ten_phong
            """, (ma_toanha,))
        
        ds_phong = con_tro.fetchall()
        
        # Kiem tra trang thai bao hong cua tung phong
        for phong in ds_phong:
            # Kiem tra phieu dang cho xac nhan (KTV da sua xong)
            con_tro.execute("""
                SELECT COUNT(*) AS so_luong FROM phieubaohong 
                WHERE ma_phong = %s AND trang_thai = 'da_hoan_thanh'
            """, (phong['ma_phong'],))
            cho_xac_nhan = con_tro.fetchone()
            
            # Kiem tra phieu dang xu ly (chua hoan thanh)
            con_tro.execute("""
                SELECT muc_do, trang_thai FROM phieubaohong 
                WHERE ma_phong = %s AND trang_thai NOT IN ('da_xac_nhan', 'khong_duyet', 'tu_choi_sua', 'da_hoan_thanh')
                ORDER BY FIELD(muc_do, 'khancap', 'thuong') 
                LIMIT 1
            """, (phong['ma_phong'],))
            bao_hong = con_tro.fetchone()
            
            if bao_hong:
                if bao_hong['muc_do'] == 'khancap':
                    phong['mau_trang_thai'] = 'do'
                else:
                    phong['mau_trang_thai'] = 'vang'
            elif cho_xac_nhan and cho_xac_nhan['so_luong'] > 0:
                phong['mau_trang_thai'] = 'tim'
            else:
                phong['mau_trang_thai'] = 'xanh'
        
        return jsonify({'thanh_cong': True, 'du_lieu': ds_phong})
    finally:
        conn.close()

@app.route('/api/quanly/phieu-phong/<int:ma_phong>', methods=['GET'])
def ql_phieu_phong(ma_phong):
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("""
            SELECT pb.*, nd.ho_ten AS ten_nguoi_bao, nd.so_dien_thoai AS sdt_nguoi_bao, p.ten_phong, t.ten_toanha,
                   pc.anh_hoan_thanh, pc.ghi_chu, kt.ho_ten AS ten_ky_thuat, kt.so_dien_thoai AS sdt_ky_thuat
            FROM phieubaohong pb
            JOIN nguoidung nd ON pb.ma_nguoidung = nd.ma_nguoidung
            JOIN phong p ON pb.ma_phong = p.ma_phong
            JOIN toanha t ON pb.ma_toanha = t.ma_toanha
            LEFT JOIN phancong pc ON pc.ma_phieu = pb.ma_phieu AND pc.trang_thai IN ('da_tiep_nhan', 'hoan_thanh')
            LEFT JOIN nguoidung kt ON pc.ma_kythuat = kt.ma_nguoidung
            WHERE pb.ma_phong = %s AND pb.trang_thai NOT IN ('da_xac_nhan', 'khong_duyet')
            ORDER BY pb.ngay_tao DESC
        """, (ma_phong,))
        ds_phieu = con_tro.fetchall()
        for phieu in ds_phieu:
            phieu['ngay_tao'] = phieu['ngay_tao'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_tao'] else ''
            phieu['ngay_cap_nhat'] = phieu['ngay_cap_nhat'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_cap_nhat'] else ''
            
        con_tro.execute("""
            SELECT ho_ten, tai_khoan as ma_sv, lop, khoa, so_dien_thoai, anh_dai_dien
            FROM nguoidung
            WHERE ma_phong = %s AND vai_tro = 'sinhvien' AND trang_thai = 'da_duyet'
        """, (ma_phong,))
        ds_sinhvien = con_tro.fetchall()
        
        return jsonify({'thanh_cong': True, 'du_lieu': ds_phieu, 'sinh_vien': ds_sinhvien})
    finally:
        conn.close()

@app.route('/api/quanly/phan-cong', methods=['POST'])
def phan_cong_ky_thuat():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    du_lieu = request.get_json()
    ma_phieu = du_lieu.get('ma_phieu')
    ma_kythuat = du_lieu.get('ma_kythuat')
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor()
        con_tro.execute("""
            INSERT INTO phancong (ma_phieu, ma_kythuat, ma_quanly, trang_thai)
            VALUES (%s, %s, %s, 'cho_tiep_nhan')
        """, (ma_phieu, ma_kythuat, session['ma_nguoidung']))
        con_tro.execute("UPDATE phieubaohong SET trang_thai = 'da_phan_cong' WHERE ma_phieu = %s", (ma_phieu,))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Phân công thành công'})
    finally:
        conn.close()

@app.route('/api/quanly/khong-duyet/<int:ma_phieu>', methods=['POST'])
def khong_duyet_phieu(ma_phieu):
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    
    du_lieu = request.get_json(silent=True) or {}
    ly_do = du_lieu.get('lydo', '')
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor()
        if ly_do:
            con_tro.execute("UPDATE phieubaohong SET trang_thai = 'khong_duyet', mo_ta = CONCAT(IFNULL(mo_ta, ''), %s) WHERE ma_phieu = %s", (f"\n[Lý do từ chối: {ly_do}]", ma_phieu,))
        else:
            con_tro.execute("UPDATE phieubaohong SET trang_thai = 'khong_duyet' WHERE ma_phieu = %s", (ma_phieu,))
        # Huy tat ca phancong lien quan (neu la phieu khan cap da tu dong phan cong)
        con_tro.execute("DELETE FROM phancong WHERE ma_phieu = %s AND trang_thai = 'cho_tiep_nhan'", (ma_phieu,))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Đã huỷ phiếu báo hỏng'})
    finally:
        conn.close()

@app.route('/api/quanly/xac-nhan/<int:ma_phieu>', methods=['POST'])
def xac_nhan_hoan_thanh(ma_phieu):
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    du_lieu = request.get_json()
    xac_nhan = du_lieu.get('xac_nhan', True)
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor()
        if xac_nhan:
            con_tro.execute("UPDATE phieubaohong SET trang_thai = 'da_xac_nhan' WHERE ma_phieu = %s", (ma_phieu,))
            con_tro.execute("""
                UPDATE phancong SET trang_thai = 'hoan_thanh' 
                WHERE ma_phieu = %s AND trang_thai = 'da_tiep_nhan'
            """, (ma_phieu,))
            thong_bao = 'Đã xác nhận hoàn thành'
        else:
            con_tro.execute("UPDATE phieubaohong SET trang_thai = 'da_phan_cong' WHERE ma_phieu = %s", (ma_phieu,))
            thong_bao = 'Đã từ chối xác nhận. Yêu cầu kỹ thuật viên sửa lại.'
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': thong_bao})
    finally:
        conn.close()

@app.route('/api/quanly/ds-cho-duyet', methods=['GET'])
def ds_cho_duyet():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("""
            SELECT nd.*, p.ten_phong, t.ten_toanha
            FROM nguoidung nd
            LEFT JOIN phong p ON nd.ma_phong = p.ma_phong
            LEFT JOIN toanha t ON p.ma_toanha = t.ma_toanha
            WHERE nd.trang_thai = 'cho_duyet' AND nd.vai_tro = 'sinhvien'
        """)
        ds = con_tro.fetchall()
        for nd in ds:
            nd.pop('mat_khau', None)
            if nd.get('ngay_tao'):
                nd['ngay_tao'] = nd['ngay_tao'].strftime('%d/%m/%Y %H:%M')
        return jsonify({'thanh_cong': True, 'du_lieu': ds})
    finally:
        conn.close()

@app.route('/api/quanly/duyet-dang-ky', methods=['POST'])
def duyet_dang_ky():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    du_lieu = request.get_json()
    ma_nguoidung = du_lieu.get('ma_nguoidung')
    chap_nhan = du_lieu.get('chap_nhan', True)
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("SELECT ho_ten, tai_khoan, email FROM nguoidung WHERE ma_nguoidung = %s", (ma_nguoidung,))
        sv = con_tro.fetchone()

        trang_thai_moi = 'da_duyet' if chap_nhan else 'tu_choi'
        con_tro.execute("UPDATE nguoidung SET trang_thai = %s WHERE ma_nguoidung = %s", (trang_thai_moi, ma_nguoidung))
        conn.commit()
        
        # Gui email
        thong_bao_kq = f'Đã {"duyệt" if chap_nhan else "từ chối"} tài khoản.'
        if chap_nhan and sv and sv.get('email'):
            tieu_de = "✅ Duyệt Tài Khoản KTX Thành Công"
            noi_dung = f'''
            <h3>Chào bạn {sv.get('ho_ten')},</h3>
            <p>Tài khoản Hệ Thống Báo Hỏng KTX của bạn đã được Quản Lý KTX phê duyệt thành công.</p>
            <p><b>Tài khoản đăng nhập của bạn (MSSV):</b> {sv.get('tai_khoan')}</p>
            <p>Bây giờ bạn đã có thể truy cập hệ thống để tạo phiếu báo hỏng tại phòng của mình.</p>
            <br>
            <p>Cảm ơn bạn.<br><b>Ban Quản Lý KTX</b></p>
            '''
            threading.Thread(target=gui_email_thong_bao, args=(sv.get('email'), tieu_de, noi_dung), daemon=True).start()
            thong_bao_kq += ' Đã gửi Email thông báo tới sinh viên.'

        return jsonify({'thanh_cong': True, 'thong_bao': thong_bao_kq})
    finally:
        conn.close()

@app.route('/api/quanly/ds-kythuat', methods=['GET'])
def ds_ky_thuat_vien():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("""
            SELECT nd.ma_nguoidung, nd.ho_ten, nd.tai_khoan,
                   (SELECT COUNT(*) FROM phancong pc 
                    JOIN phieubaohong pb ON pc.ma_phieu = pb.ma_phieu
                    WHERE pc.ma_kythuat = nd.ma_nguoidung 
                    AND pb.trang_thai IN ('da_phan_cong', 'dang_xu_ly', 'da_hoan_thanh')) AS dang_ban
            FROM nguoidung nd 
            WHERE nd.vai_tro = 'kythuat' AND nd.trang_thai = 'da_duyet'
        """)
        ds = con_tro.fetchall()
        return jsonify({'thanh_cong': True, 'du_lieu': ds})
    finally:
        conn.close()

@app.route('/api/quanly/thong-ke', methods=['GET'])
def thong_ke():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    
    loai = request.args.get('loai', 'thang')  # thang hoac nam
    nam = request.args.get('nam', datetime.now().year)
    thang = request.args.get('thang', datetime.now().month)
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        
        if loai == 'thang':
            dieu_kien = "MONTH(pb.ngay_tao) = %s AND YEAR(pb.ngay_tao) = %s"
            tham_so = (thang, nam)
        else:
            dieu_kien = "YEAR(pb.ngay_tao) = %s"
            tham_so = (nam,)
        
        # Tong so phieu
        con_tro.execute(f"SELECT COUNT(*) as tong FROM phieubaohong pb WHERE {dieu_kien}", tham_so)
        tong = con_tro.fetchone()['tong']
        
        # Theo trang thai
        con_tro.execute(f"""
            SELECT trang_thai, COUNT(*) as so_luong 
            FROM phieubaohong pb WHERE {dieu_kien}
            GROUP BY trang_thai
        """, tham_so)
        theo_trang_thai = con_tro.fetchall()
        
        # Theo muc do
        con_tro.execute(f"""
            SELECT muc_do, COUNT(*) as so_luong 
            FROM phieubaohong pb WHERE {dieu_kien}
            GROUP BY muc_do
        """, tham_so)
        theo_muc_do = con_tro.fetchall()
        
        # Danh sach phieu trong thong ke
        con_tro.execute(f"""
            SELECT pb.*, nd.ho_ten AS ten_nguoi_bao, nd.so_dien_thoai AS sdt_nguoi_bao, p.ten_phong, t.ten_toanha,
                   pc.anh_hoan_thanh, pc.ghi_chu, pc.ngay_cap_nhat AS thoi_gian_sua_xong, kt.ho_ten AS ten_ky_thuat, kt.so_dien_thoai AS sdt_ky_thuat
            FROM phieubaohong pb
            JOIN nguoidung nd ON pb.ma_nguoidung = nd.ma_nguoidung
            JOIN phong p ON pb.ma_phong = p.ma_phong
            JOIN toanha t ON pb.ma_toanha = t.ma_toanha
            LEFT JOIN phancong pc ON pc.ma_phieu = pb.ma_phieu AND pc.trang_thai IN ('da_tiep_nhan', 'hoan_thanh')
            LEFT JOIN nguoidung kt ON pc.ma_kythuat = kt.ma_nguoidung
            WHERE {dieu_kien}
            ORDER BY pb.ngay_tao DESC
        """, tham_so)
        ds_phieu = con_tro.fetchall()
        for phieu in ds_phieu:
            phieu['ngay_tao'] = phieu['ngay_tao'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_tao'] else ''
            phieu['ngay_cap_nhat'] = phieu['ngay_cap_nhat'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_cap_nhat'] else ''
            phieu['thoi_gian_sua_xong'] = phieu['thoi_gian_sua_xong'].strftime('%d/%m/%Y %H:%M') if phieu.get('thoi_gian_sua_xong') else ''
        
        return jsonify({
            'thanh_cong': True,
            'du_lieu': {
                'tong': tong,
                'theo_trang_thai': theo_trang_thai,
                'theo_muc_do': theo_muc_do,
                'danh_sach_phieu': ds_phieu
            }
        })
    finally:
        conn.close()

@app.route('/api/quanly/xuat-bao-cao', methods=['GET'])
def xuat_bao_cao():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    
    loai = request.args.get('loai', 'thang')
    nam = request.args.get('nam', datetime.now().year)
    thang = request.args.get('thang', datetime.now().month)
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        
        if loai == 'thang':
            dieu_kien = "MONTH(pb.ngay_tao) = %s AND YEAR(pb.ngay_tao) = %s"
            tham_so = (thang, nam)
            ten_file = f"BaoCao_KTX_Thang{thang}_{nam}.xlsx"
        else:
            dieu_kien = "YEAR(pb.ngay_tao) = %s"
            tham_so = (nam,)
            ten_file = f"BaoCao_KTX_Nam_{nam}.xlsx"
            
        con_tro.execute(f"""
            SELECT pb.ma_phieu AS 'Mã Phiếu',
                   pb.ten_loi AS 'Tên Lỗi',
                   pb.muc_do AS 'Mức Độ',
                   pb.trang_thai AS 'Trạng Thái',
                   p.ten_phong AS 'Phòng',
                   t.ten_toanha AS 'Tòa Nhà',
                   nd.ho_ten AS 'Người Báo',
                   nd.so_dien_thoai AS 'SĐT Người Báo',
                   pb.ngay_tao AS 'Ngày Báo',
                   kt.ho_ten AS 'Kỹ Thuật Viên',
                   pc.ngay_cap_nhat AS 'Ngày Hoàn Thành',
                   pc.ghi_chu AS 'Ghi Chú KTV'
            FROM phieubaohong pb
            JOIN nguoidung nd ON pb.ma_nguoidung = nd.ma_nguoidung
            JOIN phong p ON pb.ma_phong = p.ma_phong
            JOIN toanha t ON pb.ma_toanha = t.ma_toanha
            LEFT JOIN phancong pc ON pc.ma_phieu = pb.ma_phieu AND pc.trang_thai IN ('da_tiep_nhan', 'hoan_thanh')
            LEFT JOIN nguoidung kt ON pc.ma_kythuat = kt.ma_nguoidung
            WHERE {dieu_kien}
            ORDER BY pb.ngay_tao DESC
        """, tham_so)
        
        ds_phieu = con_tro.fetchall()
        
        # Translate states
        trang_thai_dict = {
            'cho_xu_ly': 'Chờ xử lý',
            'da_phan_cong': 'Đã phân công',
            'dang_xu_ly': 'Đang sửa chữa',
            'da_hoan_thanh': 'Chờ QL duyệt',
            'da_xac_nhan': 'Đã hoàn thành',
            'khong_duyet': 'Không duyệt',
            'tu_choi_sua': 'Từ chối sửa'
        }
        
        for phieu in ds_phieu:
            phieu['Mức Độ'] = 'Khẩn cấp' if phieu['Mức Độ'] == 'khancap' else 'Thường'
            phieu['Trạng Thái'] = trang_thai_dict.get(phieu['Trạng Thái'], phieu['Trạng Thái'])
            if phieu['Ngày Báo']:
                phieu['Ngày Báo'] = phieu['Ngày Báo'].strftime('%d/%m/%Y %H:%M:%S')
            if phieu['Ngày Hoàn Thành']:
                phieu['Ngày Hoàn Thành'] = phieu['Ngày Hoàn Thành'].strftime('%d/%m/%Y %H:%M:%S')
                
        df = pd.DataFrame(ds_phieu)
        if df.empty:
            columns = ['Mã Phiếu', 'Tên Lỗi', 'Mức Độ', 'Trạng Thái', 'Phòng', 'Tòa Nhà', 'Người Báo', 'SĐT Người Báo', 'Ngày Báo', 'Kỹ Thuật Viên', 'Ngày Hoàn Thành', 'Ghi Chú KTV']
            df = pd.DataFrame(columns=columns)
            
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='DanhSachBaoHong', startrow=2)
            workbook = writer.book
            worksheet = writer.sheets['DanhSachBaoHong']
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # 1. Thêm tiêu đề lớn ở trên cùng (Merge cells)
            tieu_de_str = f"BÁO CÁO THỐNG KÊ BÁO HỎNG KTX - {'THÁNG ' + str(thang) + '/' if loai == 'thang' else 'NĂM '}{nam}"
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
            title_cell = worksheet.cell(row=1, column=1, value=tieu_de_str)
            title_cell.font = Font(size=14, bold=True, color='000080')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Styling thiết lập chung
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 2. Chỉnh độ rộng cột
            column_widths = {
                'A': 10, 'B': 25, 'C': 12, 'D': 15, 'E': 10, 'F': 15, 
                'G': 20, 'H': 15, 'I': 20, 'J': 20, 'K': 20, 'L': 30
            }
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 3. Format Dòng Header (dòng số 3)
            for cell in worksheet[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
                
            # 4. Format Dữ Liệu
            if not df.empty:
                for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=12):
                    for idx, cell in enumerate(row):
                        cell.border = thin_border
                        # Căn giữa cho Mã phiếu, Mức độ, Trạng thái, Phòng, Toà nhà, SĐT, Ngày báo, Ngày HT
                        if idx in [0, 2, 3, 4, 5, 7, 8, 10]:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            cell.alignment = Alignment(vertical='center', wrap_text=True)

        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=ten_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    finally:
        conn.close()

@app.route('/api/quanly/lich-su', methods=['GET'])
def ql_lich_su():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("""
            SELECT pb.*, nd.ho_ten AS ten_nguoi_bao, nd.so_dien_thoai AS sdt_nguoi_bao, p.ten_phong, t.ten_toanha,
                   pc.anh_hoan_thanh, pc.ghi_chu, pc.ngay_cap_nhat AS thoi_gian_sua_xong, kt.ho_ten AS ten_ky_thuat, kt.so_dien_thoai AS sdt_ky_thuat
            FROM phieubaohong pb
            JOIN nguoidung nd ON pb.ma_nguoidung = nd.ma_nguoidung
            JOIN phong p ON pb.ma_phong = p.ma_phong
            JOIN toanha t ON pb.ma_toanha = t.ma_toanha
            LEFT JOIN phancong pc ON pc.ma_phieu = pb.ma_phieu AND pc.trang_thai IN ('da_tiep_nhan', 'hoan_thanh')
            LEFT JOIN nguoidung kt ON pc.ma_kythuat = kt.ma_nguoidung
            ORDER BY pb.ngay_tao DESC
        """)
        ds_phieu = con_tro.fetchall()
        for phieu in ds_phieu:
            phieu['ngay_tao'] = phieu['ngay_tao'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_tao'] else ''
            phieu['ngay_cap_nhat'] = phieu['ngay_cap_nhat'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_cap_nhat'] else ''
            phieu['thoi_gian_sua_xong'] = phieu['thoi_gian_sua_xong'].strftime('%d/%m/%Y %H:%M') if phieu.get('thoi_gian_sua_xong') else ''
        return jsonify({'thanh_cong': True, 'du_lieu': ds_phieu})
    finally:
        conn.close()

# ============================
# API KY THUAT VIEN
# ============================
@app.route('/api/kythuat/thong-tin', methods=['GET'])
def lay_thong_tin_kt():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("SELECT * FROM nguoidung WHERE ma_nguoidung = %s", (session['ma_nguoidung'],))
        thong_tin = con_tro.fetchone()
        if thong_tin:
            thong_tin.pop('mat_khau', None)
            if thong_tin.get('ngay_sinh'):
                thong_tin['ngay_sinh'] = thong_tin['ngay_sinh'].strftime('%Y-%m-%d')
            if thong_tin.get('ngay_tao'):
                thong_tin['ngay_tao'] = thong_tin['ngay_tao'].strftime('%Y-%m-%d %H:%M:%S')
        return jsonify({'thanh_cong': True, 'du_lieu': thong_tin})
    finally:
        conn.close()

@app.route('/api/kythuat/cap-nhat', methods=['PUT'])
def cap_nhat_kt():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    du_lieu = request.get_json()
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor()
        con_tro.execute("""
            UPDATE nguoidung SET ho_ten=%s, ngay_sinh=%s, so_dien_thoai=%s
            WHERE ma_nguoidung=%s
        """, (
            du_lieu.get('ho_ten'),
            du_lieu.get('ngay_sinh'),
            du_lieu.get('so_dien_thoai'),
            session['ma_nguoidung']
        ))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Cập nhật thông tin thành công'})
    finally:
        conn.close()

@app.route('/api/kythuat/ds-phieu', methods=['GET'])
def kt_ds_phieu():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        try:
            con_tro = conn.cursor(dictionary=True)
            con_tro.execute("""
                SELECT pc.ma_phancong, pc.ma_phieu, pc.ma_kythuat, pc.ma_quanly, pc.trang_thai AS trang_thai_phancong, pc.anh_hoan_thanh, pc.ghi_chu, pc.ngay_phan_cong, pc.ngay_cap_nhat AS ngay_cap_nhat_phancong,
                       pb.ten_loi, pb.mo_ta, pb.anh_minh_chung, pb.muc_do, pb.trang_thai AS trang_thai_phieu,
                       pb.ngay_tao AS ngay_tao_phieu, p.ten_phong, t.ten_toanha, 
                       nd.ho_ten AS ten_nguoi_bao, nd.so_dien_thoai AS sdt_nguoi_bao
                FROM phancong pc
                JOIN phieubaohong pb ON pc.ma_phieu = pb.ma_phieu
                LEFT JOIN phong p ON pb.ma_phong = p.ma_phong
                LEFT JOIN toanha t ON pb.ma_toanha = t.ma_toanha
                LEFT JOIN nguoidung nd ON pb.ma_nguoidung = nd.ma_nguoidung
                WHERE pc.ma_kythuat = %s
                ORDER BY pc.ngay_phan_cong DESC
            """, (session['ma_nguoidung'],))
            ds_phieu = con_tro.fetchall()
            for phieu in ds_phieu:
                phieu['ngay_phan_cong'] = phieu['ngay_phan_cong'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_phan_cong'] else ''
                phieu['ngay_cap_nhat_phancong'] = phieu['ngay_cap_nhat_phancong'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_cap_nhat_phancong'] else ''
                phieu['ngay_tao_phieu'] = phieu['ngay_tao_phieu'].strftime('%d/%m/%Y %H:%M') if phieu['ngay_tao_phieu'] else ''
            return jsonify({'thanh_cong': True, 'du_lieu': ds_phieu})
        except Exception as e:
            import traceback
            err = traceback.format_exc()
            print("ERROR IN DS PHIEU:", err)
            with open('log_err.txt', 'a') as f:
                f.write(err + "\n")
            return jsonify({'thanh_cong': False, 'error': str(err)}), 500
    finally:
        conn.close()

@app.route('/api/kythuat/tiep-nhan/<int:ma_phancong>', methods=['POST'])
def tiep_nhan_sua(ma_phancong):
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        # Lay ma phieu
        con_tro.execute("SELECT ma_phieu FROM phancong WHERE ma_phancong = %s", (ma_phancong,))
        pc = con_tro.fetchone()
        # Cap nhat phancong nay thanh da_tiep_nhan
        con_tro.execute("UPDATE phancong SET trang_thai = 'da_tiep_nhan' WHERE ma_phancong = %s", (ma_phancong,))
        if pc:
            # Huy tat ca phancong khac cua cung phieu (KTV khac se khong thay nua)
            con_tro.execute("""
                DELETE FROM phancong 
                WHERE ma_phieu = %s AND ma_phancong != %s
            """, (pc['ma_phieu'], ma_phancong))
            # Cap nhat trang thai phieu
            con_tro.execute("UPDATE phieubaohong SET trang_thai = 'dang_xu_ly' WHERE ma_phieu = %s", (pc['ma_phieu'],))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Đã tiếp nhận phiếu sửa chữa'})
    finally:
        conn.close()

@app.route('/api/kythuat/tu-choi/<int:ma_phancong>', methods=['POST'])
def tu_choi_sua(ma_phancong):
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        # Lay ma phieu tu phancong
        con_tro.execute("SELECT ma_phieu FROM phancong WHERE ma_phancong = %s", (ma_phancong,))
        pc = con_tro.fetchone()
        # Cap nhat phancong thanh tu_choi
        con_tro.execute("UPDATE phancong SET trang_thai = 'tu_choi' WHERE ma_phancong = %s", (ma_phancong,))
        # Huy phieu - phong tro lai binh thuong
        if pc:
            con_tro.execute("UPDATE phieubaohong SET trang_thai = 'khong_duyet' WHERE ma_phieu = %s", (pc['ma_phieu'],))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Đã từ chối. Phiếu đã bị huỷ.'})
    finally:
        conn.close()

@app.route('/api/kythuat/hoan-thanh/<int:ma_phancong>', methods=['POST'])
def hoan_thanh_sua(ma_phancong):
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    
    # Xu ly upload anh hoan thanh
    ten_file_anh = ''
    if 'anh_hoan_thanh' in request.files:
        file_anh = request.files['anh_hoan_thanh']
        if file_anh and file_anh.filename and kiem_tra_duoi_file(file_anh.filename):
            if os.environ.get('CLOUDINARY_CLOUD_NAME'):
                try:
                    upload_result = cloudinary.uploader.upload(file_anh)
                    ten_file_anh = upload_result.get('secure_url')
                except Exception as e:
                    print(f"Loi upload Cloudinary: {e}")
            else:
                ten_file = secure_filename(f"ht_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file_anh.filename}")
                file_anh.save(os.path.join(app.config['THU_MUC_UPLOAD'], ten_file))
                ten_file_anh = ten_file
    
    ghi_chu = request.form.get('ghi_chu', '')
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("""
            UPDATE phancong SET trang_thai = 'hoan_thanh', anh_hoan_thanh = %s, ghi_chu = %s
            WHERE ma_phancong = %s
        """, (ten_file_anh, ghi_chu, ma_phancong))
        
        # Cap nhat trang thai phieu
        con_tro.execute("SELECT ma_phieu FROM phancong WHERE ma_phancong = %s", (ma_phancong,))
        pc = con_tro.fetchone()
        if pc:
            con_tro.execute("UPDATE phieubaohong SET trang_thai = 'da_hoan_thanh' WHERE ma_phieu = %s", (pc['ma_phieu'],))
        
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Đã gửi minh chứng hoàn thành'})
    finally:
        conn.close()

# ============================
# API ADMIN
# ============================
@app.route('/api/admin/ds-nguoidung', methods=['GET'])
def admin_ds_nguoidung():
    if 'ma_nguoidung' not in session or session.get('vai_tro') != 'admin':
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("""
            SELECT nd.ma_nguoidung, nd.tai_khoan, nd.ho_ten, nd.vai_tro, nd.trang_thai, 
                   nd.so_dien_thoai, nd.ngay_tao, t.ten_toanha
            FROM nguoidung nd
            LEFT JOIN toanha t ON nd.ma_toanha_quanly = t.ma_toanha
            WHERE nd.vai_tro != 'admin'
            ORDER BY nd.vai_tro, nd.ngay_tao DESC
        """)
        ds = con_tro.fetchall()
        for nd in ds:
            if nd.get('ngay_tao'):
                nd['ngay_tao'] = nd['ngay_tao'].strftime('%d/%m/%Y %H:%M')
        return jsonify({'thanh_cong': True, 'du_lieu': ds})
    finally:
        conn.close()

@app.route('/api/admin/tao-tai-khoan', methods=['POST'])
def admin_tao_tai_khoan():
    if 'ma_nguoidung' not in session or session.get('vai_tro') != 'admin':
        return jsonify({'thanh_cong': False}), 401
    
    du_lieu = request.get_json()
    tai_khoan = du_lieu.get('tai_khoan', '')
    mat_khau = du_lieu.get('mat_khau', '')
    ho_ten = du_lieu.get('ho_ten', '')
    vai_tro = du_lieu.get('vai_tro', '')
    ma_toanha_quanly = du_lieu.get('ma_toanha_quanly')
    so_dien_thoai = du_lieu.get('so_dien_thoai', '')
    
    if not all([tai_khoan, mat_khau, ho_ten, vai_tro]):
        return jsonify({'thanh_cong': False, 'thong_bao': 'Vui lòng điền đầy đủ thông tin'})
    
    if vai_tro not in ['quanly', 'kythuat']:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Vai trò không hợp lệ'})
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("SELECT * FROM nguoidung WHERE tai_khoan = %s", (tai_khoan,))
        if con_tro.fetchone():
            return jsonify({'thanh_cong': False, 'thong_bao': 'Tài khoản đã tồn tại'})
        
        hashed_password = generate_password_hash(mat_khau)
        con_tro.execute("""
            INSERT INTO nguoidung (tai_khoan, mat_khau, ho_ten, vai_tro, trang_thai, ma_toanha_quanly, so_dien_thoai)
            VALUES (%s, %s, %s, %s, 'da_duyet', %s, %s)
        """, (tai_khoan, hashed_password, ho_ten, vai_tro, ma_toanha_quanly if vai_tro == 'quanly' else None, so_dien_thoai))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': f'Tạo tài khoản {vai_tro} thành công'})
    finally:
        conn.close()

@app.route('/api/admin/xoa-tai-khoan/<int:ma_nguoidung>', methods=['DELETE'])
def admin_xoa_tai_khoan(ma_nguoidung):
    if 'ma_nguoidung' not in session or session.get('vai_tro') != 'admin':
        return jsonify({'thanh_cong': False}), 401
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor()
        con_tro.execute("DELETE FROM nguoidung WHERE ma_nguoidung = %s AND vai_tro != 'admin'", (ma_nguoidung,))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Đã xoá tài khoản'})
    finally:
        conn.close()

# ============================
# API THONG BAO KHAN CAP
# ============================
@app.route('/api/thong-bao/khan-cap', methods=['GET'])
def thong_bao_khan_cap():
    """Ky thuat vien va quan ly polling de kiem tra phieu khan cap moi"""
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False}), 401
    
    sau_ma_phieu = request.args.get('sau_ma_phieu', 0, type=int)
    vai_tro = session.get('vai_tro')
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False}), 500
    try:
        con_tro = conn.cursor(dictionary=True)
        
        if vai_tro == 'kythuat':
            # Ky thuat vien: kiem tra phieu khan cap moi duoc phan cong
            con_tro.execute("""
                SELECT pc.ma_phancong, pb.ma_phieu, pb.ten_loi, pb.mo_ta, pb.muc_do,
                       p.ten_phong, t.ten_toanha, nd.ho_ten AS ten_nguoi_bao, nd.so_dien_thoai AS sdt_nguoi_bao,
                       pb.ngay_tao
                FROM phancong pc
                JOIN phieubaohong pb ON pc.ma_phieu = pb.ma_phieu
                JOIN phong p ON pb.ma_phong = p.ma_phong
                JOIN toanha t ON pb.ma_toanha = t.ma_toanha
                JOIN nguoidung nd ON pb.ma_nguoidung = nd.ma_nguoidung
                WHERE pc.ma_kythuat = %s 
                AND pb.muc_do = 'khancap'
                AND pc.trang_thai = 'cho_tiep_nhan'
                AND pb.ma_phieu > %s
                ORDER BY pb.ngay_tao DESC
            """, (session['ma_nguoidung'], sau_ma_phieu))
        elif vai_tro == 'quanly':
            # Quan ly: kiem tra phieu khan cap moi
            con_tro.execute("SELECT ma_toanha_quanly FROM nguoidung WHERE ma_nguoidung = %s", (session['ma_nguoidung'],))
            ql = con_tro.fetchone()
            ma_toanha = ql.get('ma_toanha_quanly') if ql else None
            
            if ma_toanha:
                # Co toa quan ly -> chi hien phieu cua toa do
                con_tro.execute("""
                    SELECT pb.ma_phieu, pb.ten_loi, pb.mo_ta, pb.muc_do,
                           p.ten_phong, t.ten_toanha, nd.ho_ten AS ten_nguoi_bao, nd.so_dien_thoai AS sdt_nguoi_bao,
                           pb.ngay_tao
                    FROM phieubaohong pb
                    JOIN phong p ON pb.ma_phong = p.ma_phong
                    JOIN toanha t ON pb.ma_toanha = t.ma_toanha
                    JOIN nguoidung nd ON pb.ma_nguoidung = nd.ma_nguoidung
                    WHERE pb.muc_do = 'khancap'
                    AND pb.ma_toanha = %s
                    AND pb.ma_phieu > %s
                    AND pb.trang_thai IN ('cho_xu_ly', 'da_phan_cong', 'dang_xu_ly')
                    ORDER BY pb.ngay_tao DESC
                """, (ma_toanha, sau_ma_phieu))
            else:
                # Chua co toa -> hien tat ca phieu khan cap
                con_tro.execute("""
                    SELECT pb.ma_phieu, pb.ten_loi, pb.mo_ta, pb.muc_do,
                           p.ten_phong, t.ten_toanha, nd.ho_ten AS ten_nguoi_bao, nd.so_dien_thoai AS sdt_nguoi_bao,
                           pb.ngay_tao
                    FROM phieubaohong pb
                    JOIN phong p ON pb.ma_phong = p.ma_phong
                    JOIN toanha t ON pb.ma_toanha = t.ma_toanha
                    JOIN nguoidung nd ON pb.ma_nguoidung = nd.ma_nguoidung
                    WHERE pb.muc_do = 'khancap'
                    AND pb.ma_phieu > %s
                    AND pb.trang_thai IN ('cho_xu_ly', 'da_phan_cong', 'dang_xu_ly')
                    ORDER BY pb.ngay_tao DESC
                """, (sau_ma_phieu,))
        else:
            return jsonify({'thanh_cong': True, 'du_lieu': []})
        
        ds_khan_cap = con_tro.fetchall()
        for phieu in ds_khan_cap:
            if phieu.get('ngay_tao'):
                phieu['ngay_tao'] = phieu['ngay_tao'].strftime('%d/%m/%Y %H:%M')
        
        return jsonify({'thanh_cong': True, 'du_lieu': ds_khan_cap})
    finally:
        conn.close()

@app.route('/api/thong-bao/dem-so-luong', methods=['GET'])
def dem_so_luong_thong_bao():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False})
    
    vai_tro = session.get('vai_tro')
    ma_nguoidung = session.get('ma_nguoidung')
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False})
    
    try:
        con_tro = conn.cursor(dictionary=True)
        ket_qua = {}
        
        if vai_tro == 'quanly':
            # Dem sv cho duyet
            con_tro.execute("SELECT COUNT(*) as cnt FROM nguoidung WHERE vai_tro='sinhvien' AND trang_thai='cho_duyet'")
            ket_qua['sv_cho_duyet'] = con_tro.fetchone()['cnt']
            
            # Dem phieu can quan ly xu ly (cho_xu_ly hoac da_hoan_thanh -> can QL xac nhan)
            con_tro.execute("SELECT ma_toanha_quanly FROM nguoidung WHERE ma_nguoidung = %s", (ma_nguoidung,))
            ql = con_tro.fetchone()
            ma_toanha = ql.get('ma_toanha_quanly') if ql else None
            
            if ma_toanha:
                con_tro.execute("SELECT COUNT(*) as cnt FROM phieubaohong WHERE ma_toanha = %s AND trang_thai IN ('cho_xu_ly', 'da_hoan_thanh')", (ma_toanha,))
            else:
                con_tro.execute("SELECT COUNT(*) as cnt FROM phieubaohong WHERE trang_thai IN ('cho_xu_ly', 'da_hoan_thanh')")
            ket_qua['phieu_can_xu_ly'] = con_tro.fetchone()['cnt']
            
        elif vai_tro == 'sinhvien':
            con_tro.execute("SELECT COUNT(*) as cnt FROM phieubaohong WHERE ma_phong = (SELECT ma_phong FROM nguoidung WHERE ma_nguoidung = %s) AND trang_thai IN ('da_phan_cong', 'dang_xu_ly', 'da_hoan_thanh')", (ma_nguoidung,))
            ket_qua['phieu_dang_xu_ly'] = con_tro.fetchone()['cnt']
            
        elif vai_tro == 'kythuat':
            con_tro.execute("SELECT COUNT(*) as cnt FROM phancong WHERE ma_kythuat = %s AND trang_thai = 'cho_tiep_nhan'", (ma_nguoidung,))
            ket_qua['phieu_cho_tiep_nhan'] = con_tro.fetchone()['cnt']
            
        return jsonify({'thanh_cong': True, 'du_lieu': ket_qua})
    finally:
        conn.close()

# ============================
# CAC HAM GUI EMAIL VÀ QUEN MAT KHAU
# ============================
def gui_email_thong_bao(email_nhan, tieu_de, noi_dung):
    import time
    import os
    email_gui = os.environ.get('MAIL_USERNAME', '').strip()
    mat_khau = os.environ.get('MAIL_PASSWORD', '').strip()
    
    log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'email_log.txt')
    with open(log_path, 'a', encoding='utf-8') as f:
        ghilog = lambda msg: f.write(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}\n")
        
        # Tạm thời cứ thử gửi nếu có setup trong .env, nếu không có thì bỏ qua
        if not email_gui or not mat_khau:
            ghilog(f"BỎ QUA: Chưa thiết lập MAIL_USERNAME và MAIL_PASSWORD. Không gửi được tới: {email_nhan}")
            return False
            
        try:
            msg = MIMEMultipart()
            msg['From'] = email_gui
            msg['To'] = email_nhan
            msg['Subject'] = tieu_de
            msg.attach(MIMEText(noi_dung, 'html', 'utf-8'))
            
            ghilog(f"BAT DAU GUI: {email_nhan} | Tieu de: {tieu_de}")
            
            # Ghi đề tạm thời getaddrinfo để ép hệ thống chỉ tìm và gài IPv4
            import socket
            old_getaddrinfo = socket.getaddrinfo
            def ipv4_getaddrinfo(host, port, family=0, *args, **kwargs):
                return old_getaddrinfo(host, port, socket.AF_INET, *args, **kwargs)
            socket.getaddrinfo = ipv4_getaddrinfo
            
            try:
                server = smtplib.SMTP('smtp.gmail.com', 587, timeout=15)
                server.set_debuglevel(0)
                server.starttls()
                server.login(email_gui, mat_khau)
                server.send_message(msg)
                server.quit()
            finally:
                # Phục hồi getaddrinfo
                socket.getaddrinfo = old_getaddrinfo
                
            ghilog(f"THANH CONG: Đã gửi email tới {email_nhan}")
            return True
        except Exception as e:
            ghilog(f"LOI GUI EMAIL tới {email_nhan}: {str(e)}")
            return False

@app.route('/debug/logs')
def xem_log():
    try:
        import os
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'email_log.txt')
        with open(log_path, 'r', encoding='utf-8') as f:
            return f.read(), 200, {'Content-Type': 'text/plain; charset=utf-8'}
    except Exception as e:
        return f"Không thể đọc file: {str(e)}", 500

@app.route('/api/quen-mat-khau', methods=['POST'])
def quen_mat_khau():
    du_lieu = request.get_json()
    vai_tro = du_lieu.get('vai_tro')
    tai_khoan = du_lieu.get('tai_khoan')
    xac_minh = du_lieu.get('xac_minh')
    
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Lỗi server'}), 500
        
    try:
        con_tro = conn.cursor(dictionary=True)
        if vai_tro == 'sinhvien':
            con_tro.execute("SELECT * FROM nguoidung WHERE tai_khoan=%s AND email=%s AND vai_tro='sinhvien'", (tai_khoan, xac_minh))
            sv = con_tro.fetchone()
            if not sv:
                return jsonify({'thanh_cong': False, 'thong_bao': 'MSSV hoặc Email không khớp!'})
                
            # Tao mat khau ngau nhien cho sinh vien
            mk_moi = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
            hashed_mk = generate_password_hash(mk_moi)
            
            con_tro.execute("UPDATE nguoidung SET mat_khau=%s WHERE ma_nguoidung=%s", (hashed_mk, sv['ma_nguoidung']))
            conn.commit()
            
            # Gui email
            tieu_de = "🔑 Đặt Lại Mật Khẩu KTX Thành Công"
            noi_dung = f'''
            <h3>Chào bạn {sv['ho_ten']},</h3>
            <p>Hệ thống Ký túc xá đã thiết lập lại mật khẩu cho tài khoản <b>{tai_khoan}</b> của bạn.</p>
            <p>Mật khẩu mới của bạn là: <strong style="color:red; font-size: 18px;">{mk_moi}</strong></p>
            <p>Vui lòng đăng nhập bằng mật khẩu này và bạn có thể thay đổi sau.</p>
            '''
            gui_thanh_cong = gui_email_thong_bao(xac_minh, tieu_de, noi_dung)
            if gui_thanh_cong:
                return jsonify({'thanh_cong': True, 'thong_bao': 'Mật khẩu mới đã được gửi vào Email của bạn!'})
            else:
                return jsonify({'thanh_cong': True, 'thong_bao': 'Đã đặt lại mật khẩu nhưng tính năng gửi Email chưa được bật trên máy chủ.'})
                
        else:
            # Quan ly hoac Ky thuat => Xac minh bang SĐT
            con_tro.execute("SELECT * FROM nguoidung WHERE tai_khoan=%s AND so_dien_thoai=%s AND vai_tro IN ('quanly', 'kythuat')", (tai_khoan, xac_minh))
            nv = con_tro.fetchone()
            if not nv:
                return jsonify({'thanh_cong': False, 'thong_bao': 'Tài khoản hoặc Số điện thoại không đúng!'})
            return jsonify({'thanh_cong': True, 'thong_bao': 'Xác thực thành công số điện thoại. Hãy đặt mật khẩu.'})
            
    finally:
        conn.close()

@app.route('/api/quen-mat-khau/dat-lai', methods=['POST'])
def quen_mat_khau_dat_lai():
    du_lieu = request.get_json()
    vai_tro = du_lieu.get('vai_tro')
    tai_khoan = du_lieu.get('tai_khoan')
    xac_minh = du_lieu.get('xac_minh')
    mat_khau_moi = du_lieu.get('mat_khau')
    
    if vai_tro == 'sinhvien':
        return jsonify({'thanh_cong': False, 'thong_bao': 'Không hợp lệ'}), 400
        
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Lỗi server'}), 500
        
    try:
        con_tro = conn.cursor(dictionary=True)
        # Verify lai lan nua
        con_tro.execute("SELECT * FROM nguoidung WHERE tai_khoan=%s AND so_dien_thoai=%s AND vai_tro IN ('quanly', 'kythuat')", (tai_khoan, xac_minh))
        nv = con_tro.fetchone()
        if not nv:
            return jsonify({'thanh_cong': False, 'thong_bao': 'Thông tin lỗi!'})
            
        hashed_mk = generate_password_hash(mat_khau_moi)
        con_tro.execute("UPDATE nguoidung SET mat_khau=%s WHERE ma_nguoidung=%s", (hashed_mk, nv['ma_nguoidung']))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Đã đổi mật khẩu thành công! Hãy đăng nhập lại.'})
    finally:
        conn.close()

@app.route('/api/doi-mat-khau', methods=['POST'])
def doi_mat_khau():
    if 'ma_nguoidung' not in session:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Chưa đăng nhập'}), 401
        
    du_lieu = request.get_json()
    mat_khau_cu = du_lieu.get('mat_khau_cu')
    mat_khau_moi = du_lieu.get('mat_khau_moi')
    
    if not mat_khau_cu or not mat_khau_moi:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Vui lòng điền đủ thông tin'})
        
    conn = ket_noi_db()
    if not conn:
        return jsonify({'thanh_cong': False, 'thong_bao': 'Lỗi máy chủ'}), 500
        
    try:
        con_tro = conn.cursor(dictionary=True)
        con_tro.execute("SELECT mat_khau FROM nguoidung WHERE ma_nguoidung = %s", (session['ma_nguoidung'],))
        nd = con_tro.fetchone()
        
        if not nd or not check_password_hash(nd['mat_khau'], mat_khau_cu):
            return jsonify({'thanh_cong': False, 'thong_bao': 'Mật khẩu cũ không chính xác!'})
            
        hashed_mk = generate_password_hash(mat_khau_moi)
        con_tro.execute("UPDATE nguoidung SET mat_khau = %s WHERE ma_nguoidung = %s", (hashed_mk, session['ma_nguoidung']))
        conn.commit()
        return jsonify({'thanh_cong': True, 'thong_bao': 'Đổi mật khẩu thành công!'})
    except Exception as e:
        print(f"Lỗi đổi mk: {e}")
        return jsonify({'thanh_cong': False, 'thong_bao': 'Lỗi hệ thống, thử lại sau.'})
    finally:
        conn.close()

# ============================
# CHAY UNG DUNG - Auto Reload Triggered
# ============================
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
